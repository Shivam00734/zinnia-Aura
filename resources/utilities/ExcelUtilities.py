import openpyxl
import pandas as pd
from robot.api.deco import library, keyword
from openpyxl.reader.excel import load_workbook

from resources.utilities.FileUtils import FileUtils
from resources.utilities.ReadConfig import ReadConfig


@library
class ExcelUtilities:
    def __init__(self):
        self.read_config = ReadConfig()
    file = FileUtils()

    @keyword
    def getValueFromExcel(self, excel_sheet_path, sheet_index, testcaseId, cell_number):
        value = None
        try:
            file = openpyxl.load_workbook(excel_sheet_path)
            sheet = file.worksheets[sheet_index]

            for row in sheet.iter_rows():
                cell = row[0]
                if cell.value == testcaseId:
                    valueCell = row[cell_number - 1]
                    if valueCell.value is not None:
                        value = valueCell.value
                        break

            file.close()
        except Exception as e:
            print(e)
            raise Exception("Error reading Excel file: " + str(e))
        return value

    @keyword
    def write_value_to_excel(self, excel_sheet_path, sheet_index, testcase_id, cell_number, value):
        try:
            workbook = openpyxl.load_workbook(excel_sheet_path)
            sheet = workbook.worksheets[sheet_index]

            max_column = max(sheet.max_column, 20)
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                if row[0].value == testcase_id:
                    if 1 <= cell_number <= max_column:
                        value_cell = row[cell_number - 1]
                        if value_cell.value is None:
                            value_cell.value = value
                        else:
                            value_cell.value = str(value)
                        workbook.save(excel_sheet_path)
                        workbook.close()
                        return value
                    else:
                        raise IndexError(f"Cell number {cell_number} out of range for row {row_idx}")

            raise ValueError(f"Testcase ID '{testcase_id}' not found in sheet")
        except Exception as e:
            print(e)
            raise Exception("Error reading Excel file. Close the file and try again : " + str(e))

    @keyword
    def get_test_case_ids_from_excel(self, file_path, sheetIndex):
        test_case_ids = []
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.worksheets[sheetIndex]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                test_case_id = row[1 - 1]
                if test_case_id:
                    test_case_ids.append(test_case_id)

            workbook.close()
        except Exception as e:
            print("Error reading Excel file:", e)
            raise
        return test_case_ids

    @keyword
    def check_attachment_required(self, testcase_id):
        read_config = ReadConfig()
        input_file_path = read_config.getValueByKey('ZinniaLiveTestDataND')
        # status_attachment_required = self.getValueFromExcel(input_file_path, 0, testcase_id, 8)
        status_attachment_required  = self.get_value_from_excel_by_column_name(input_file_path, "output_data", testcase_id, "${attachment_required}")

        return status_attachment_required

    @keyword
    def get_test_case_ids_from_zinnia_live_sheet(self):
        read_config = ReadConfig()
        input_file_path = read_config.getValueByKey('ZinniaLiveTestDataND')
        print(input_file_path)
        return self.get_test_case_ids_from_excel(input_file_path, 1)

    def getExcelSheetPath(self):
        read_config = ReadConfig()
        excel_file_path = read_config.getValueByKey('ZinniaLiveTestDataND')
        return excel_file_path

    @keyword
    def get_excel_testdata_by_groupId(self, group_id, sheet_name):
        read_config = ReadConfig()
        excel_file_path = read_config.getValueByKey('ZinniaLiveTestDataND')
        try:
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, dtype=str)
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None, None
        filtered_df = df[df['Group'] == group_id]
        filtered_data = filtered_df.to_numpy()

        test_case_list = []
        for idx, row in enumerate(filtered_data):
            testcase_id = str(filtered_data[idx][0])
            test_case_list.append(testcase_id)

        if not test_case_list:
            raise ValueError(f"No test cases found for group_id: {group_id}")

        return test_case_list

    @keyword
    def get_expected_data_by_test_case_id(self, test_case_id):
        import pandas as pd
        
        read_config = ReadConfig()
        excel_file_path = read_config.getValueByKey('ZinniaLiveTestDataND')
        df = pd.read_excel(excel_file_path, sheet_name="expected_result", dtype=str, keep_default_na=False)

        row = df[df['test_case_id'] == test_case_id]

        if not row.empty:
            return row.to_dict(orient='records')[0]
        else:
            raise ValueError(f"No matching test_case_id: '{test_case_id}' found in the file.")

    @keyword
    def get_value_from_excel_by_column_name(self, excel_sheet_path, sheet_name, testcase_id, column_name):
        value = None
        try:
            workbook = openpyxl.load_workbook(excel_sheet_path)
            sheet = workbook[sheet_name]

            header = [cell.value for cell in sheet[1]]
            if column_name not in header:
                raise ValueError(f"Column '{column_name}' not found in the header row")

            column_index = header.index(column_name)

            for row in sheet.iter_rows(min_row=2):
                if str(row[0].value).strip() == str(testcase_id).strip():
                    value = row[column_index].value
                    break

            workbook.close()

        except Exception as e:
            print(e)
            raise Exception("Error reading Excel file: " + str(e))

        return value


    @keyword
    def write_value_to_excel_by_column_name(self, excel_sheet_path, sheet_name, testcase_id, column_name, value):
        try:
            workbook = openpyxl.load_workbook(excel_sheet_path)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet with name '{sheet_name}' not found in the workbook")
            sheet = workbook[sheet_name]
            header_row = [cell.value for cell in sheet[1]]
            if column_name not in header_row:
                raise ValueError(f"Column name '{column_name}' not found in header row")
            column_index = header_row.index(column_name) + 1
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                if row[0].value == testcase_id:
                    if 1 <= column_index <= sheet.max_column:
                        value_cell = row[column_index - 1]
                        if value_cell.value is None:
                            value_cell.value = value
                        else:
                            value_cell.value = str(value)
                        workbook.save(excel_sheet_path)
                        workbook.close()
                        return value
                    else:
                        raise IndexError(f"Column index {column_index} out of range for row {row_idx}")
            raise ValueError(f"Testcase ID '{testcase_id}' not found in sheet")
        except Exception as e:
            print(e)
            raise Exception("Error reading Excel file. Close the file and try again: " + str(e))

    @keyword
    def clear_sheet(self):
        read_config = ReadConfig()
        excel_path = read_config.getValueByKey('ZinniaLiveTestDataND')
        workbook = load_workbook(excel_path)
        sheets_to_clear = ["output_data"]
        for sheet_name in sheets_to_clear:
            sheet = workbook[sheet_name]
            sheet.delete_rows(2, sheet.max_row - 1)
        workbook.save(excel_path)

    @keyword
    def filter_test_date_by_execution_flag(self):
        sheet_path = self.read_config.getValueByKey('transaction_onboarding_sheet')
        input_sheet_name = "Input"
        ouput_sheet_name = "output"

        df_input = pd.read_excel(sheet_path, sheet_name=input_sheet_name)

        df_filtered = df_input[df_input['${execution_flag}'] == 'YES']

        wb = load_workbook(sheet_path)
        ws = wb[ouput_sheet_name]

        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)

        for r_idx, row in enumerate(df_filtered.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Step 6: Save workbook
        wb.save(sheet_path)
        print(f"Filtered data (execution_flag == 'YES') written below existing headers in '{ouput_sheet_name}'.")



# obj = ExcelUtilities()
# obj.filter_test_date_by_execution_flag()

