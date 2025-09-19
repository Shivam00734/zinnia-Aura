from openpyxl import load_workbook
from collections import namedtuple

class ExcelLibrary:
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    
    def get_excel_data_as_list(self, file_path, sheet_name='Sheet1'):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        Row = namedtuple('Row', headers)
        
        data = []
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            data.append(Row(*values))
        return data